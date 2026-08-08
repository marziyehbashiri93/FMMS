from __future__ import annotations

import re
import shutil
import tempfile
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "019fa399-9480-7bb3-8cdd-aa2ad11dc8f0"
OUTPUT_FILE = OUTPUT_DIR / "FMMS_SAP_Workflow_Mapping.xlsx"

NAVY = "17324D"
TEAL = "087E8B"
BLUE = "DDEBF7"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"
RED = "F4CCCC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
DARK = "1F2937"
LIGHT_BORDER = Side(style="thin", color="D9E2F3")


PROCESS_HEADERS = [
    "ردیف",
    "مسیر",
    "مرحله",
    "نقش",
    "اقدام کاربر / سیستم",
    "API داخلی FMMS",
    "تغییر وضعیت / خروجی",
    "خواندن مستقیم از SAP؟",
    "OData سرویس / Entity",
    "Write به SAP؟",
    "BAPI / Function Module",
    "محل ذخیره در FMMS",
    "فیلد مرجع سند SAP",
    "محل ذخیره در SAP",
    "وضعیت پیاده‌سازی",
    "ابهام / اقدام لازم",
    "منبع کد",
]


PROCESS_ROWS = [
    [
        1,
        "آماده‌سازی داده مرجع",
        "اجرای همگام‌سازی کلی SAP",
        "مدیر سیستم / زمان‌بند",
        "اجرای دستی یا زمان‌بندی‌شده Sync برای خودرو، راننده، چک‌لیست، کاتالوگ خرابی و موجودی مرکزی",
        "POST /api/v1/sap-sync/",
        "ایجاد رکورد اجرای Sync و چهار زیرمرحله",
        "بله",
        "چهار OData فعال؛ جزئیات در شیت خواندن از SAP",
        "خیر",
        "—",
        "sap_sync_run; sap_sync_run_item",
        "—",
        "فقط Read از SAP",
        "فعال",
        "نام Celery task هنوز vehicle sync است ولی کل Sync را اجرا می‌کند.",
        "apps/integration/application/services/run_sap_sync_service.py; interfaces/api/v1/integration/views.py",
    ],
    [
        2,
        "آماده‌سازی داده مرجع",
        "خواندن خودرو و تخصیص راننده",
        "سیستم",
        "دریافت خودروها و راننده‌های اول/دوم و به‌روزرسانی snapshot تخصیص",
        "درون SAP Sync",
        "Upsert خودرو و راننده؛ ثبت تاریخچه تخصیص",
        "بله",
        "ZC_VEHICLEDRIVER_CDS / ZC_VehicleDriver",
        "خیر",
        "—",
        "vehicle; driver; vehicle_driver_assignment_history",
        "—",
        "SAP Equipment/Vehicle-Driver CDS",
        "فعال",
        "Read واقعی با SAP_USE_MOCK=False قابل اجراست.",
        "infrastructure/sap/adapters/odata/vehicle_driver_odata_adapter.py; apps/vehicle/application/services/sync_vehicles_from_sap_service.py",
    ],
    [
        3,
        "آماده‌سازی داده مرجع",
        "خواندن قالب چک‌لیست",
        "سیستم",
        "دریافت کدگروه و کد اجزای بازرسی و فعال/غیرفعال‌سازی رکوردها",
        "درون SAP Sync",
        "Upsert قالب چک‌لیست",
        "بله",
        "ZI_FLEET_CAT_B_CDS / root feed",
        "خیر",
        "—",
        "inspection_template",
        "—",
        "SAP Object Part Catalog CDS",
        "فعال",
        "Entity set پیش‌فرض خالی است و feed ریشه خوانده می‌شود.",
        "infrastructure/sap/adapters/odata/object_part_catalog_odata_adapter.py; apps/inspection/application/services/sync_inspection_templates_from_sap_service.py",
    ],
    [
        4,
        "آماده‌سازی داده مرجع",
        "خواندن کاتالوگ خرابی",
        "سیستم",
        "دریافت کد خرابی، گروه و کلاس خرابی",
        "درون SAP Sync",
        "Upsert کاتالوگ خرابی",
        "بله",
        "ZI_B_DEFECTCATALOG9_CDS / root feed",
        "خیر",
        "—",
        "fault_catalog",
        "—",
        "SAP Defect Catalog CDS",
        "فعال",
        "Entity set پیش‌فرض خالی است.",
        "infrastructure/sap/adapters/odata/fault_catalog_odata_adapter.py; apps/fault/application/services/sync_fault_catalog_from_sap_service.py",
    ],
    [
        5,
        "آماده‌سازی داده مرجع",
        "خواندن موجودی انبار مرکزی",
        "سیستم",
        "دریافت موجودی SLoc مرکزی، مقدار و ارزش موجودی",
        "درون SAP Sync",
        "Upsert موجودی مرکزی و غیرفعال‌سازی موارد حذف‌شده",
        "بله",
        "ZI_STOCK_KH08_CDS / root feed",
        "خیر",
        "—",
        "central_stock",
        "—",
        "SAP Central Stock CDS (KH08)",
        "فعال",
        "این snapshot محلی مبنای تصمیم موجودی برنامه است.",
        "infrastructure/sap/adapters/odata/central_stock_odata_adapter.py; apps/material/application/services/sync_central_stock_from_sap_service.py",
    ],
    [
        6,
        "ورود و مشاهده",
        "ورود کاربر",
        "همه نقش‌ها",
        "ورود با نام کاربری و رمز و دریافت JWT",
        "POST /api/v1/auth/token/",
        "توکن در مرورگر؛ پروفایل از FMMS",
        "خیر",
        "—",
        "خیر",
        "—",
        "fmms_users؛ tokenها در localStorage مرورگر",
        "—",
        "—",
        "فعال",
        "توکن دسترسی و refresh در دیتابیس FMMS ذخیره نمی‌شوند.",
        "frontend/src/api/client.ts; interfaces/api/v1/auth/views.py; apps/authentication/infrastructure/models.py",
    ],
    [
        7,
        "ورود و مشاهده",
        "مشاهده داشبورد و فهرست‌ها",
        "مدیر / سرپرست",
        "مشاهده KPI خودرو، راننده، خرابی، تعمیر و وضعیت تراکنش SAP",
        "GET /api/v1/vehicles/summary/ و سایر list APIs",
        "Read از جداول محلی",
        "خیر",
        "داده از snapshotهای محلی خوانده می‌شود",
        "خیر",
        "—",
        "vehicle; driver; fault; repair_order; sap_transaction",
        "—",
        "—",
        "فعال",
        "این مرحله SAP را مستقیم صدا نمی‌زند.",
        "frontend/src/features/dashboard/DashboardPage.tsx; frontend/src/api/client.ts",
    ],
    [
        8,
        "راننده",
        "انتخاب خودرو و راننده تخصیص‌یافته",
        "راننده",
        "مشاهده خودرو و رانندگان تخصیص‌یافته که قبلاً از SAP Sync شده‌اند",
        "GET /api/v1/vehicles/; GET /api/v1/drivers/",
        "نمایش snapshot محلی",
        "خیر",
        "منبع اصلی: ZC_VEHICLEDRIVER_CDS در Sync قبلی",
        "خیر",
        "—",
        "vehicle; driver; vehicle_driver_assignment_history",
        "—",
        "—",
        "فعال",
        "برای داده لحظه‌ای باید ابتدا Sync اجرا شده باشد.",
        "apps/vehicle/infrastructure/models.py; apps/driver/infrastructure/models.py",
    ],
    [
        9,
        "راننده",
        "ثبت کیلومتر",
        "راننده",
        "ثبت کیلومتر فعلی خودرو",
        "POST /api/v1/vehicles/{id}/odometer/",
        "رکورد کیلومتر با source=DRIVER",
        "خیر",
        "—",
        "خیر",
        "—",
        "vehicle_odometer_reading",
        "—",
        "—",
        "فعال",
        "در این endpoint کیلومتر به SAP نوشته نمی‌شود؛ ارسال SAP فقط هنگام ثبت خرابی انجام می‌شود.",
        "apps/vehicle/application/services/record_odometer_service.py; interfaces/api/v1/vehicle/views.py",
    ],
    [
        10,
        "راننده",
        "ایجاد چک‌لیست روزانه",
        "راننده",
        "انتخاب قالب‌ها و ثبت نوع بازرسی، کیلومتر، تاریخ و نتایج اولیه",
        "POST /api/v1/inspections/",
        "Inspection در وضعیت اولیه",
        "خیر",
        "قالب از inspection_template محلی (Sync شده از SAP)",
        "خیر",
        "—",
        "inspection; inspection_item",
        "—",
        "—",
        "فعال",
        "Read مستقیم SAP ندارد.",
        "apps/inspection/application/services/create_inspection_service.py; interfaces/api/v1/inspection/views.py",
    ],
    [
        11,
        "راننده",
        "افزودن آیتم چک‌لیست",
        "راننده",
        "ثبت OK/FAIL، توضیح و شدت هر آیتم",
        "POST /api/v1/inspections/{id}/items/",
        "افزودن/به‌روزرسانی آیتم",
        "خیر",
        "—",
        "خیر",
        "—",
        "inspection_item",
        "—",
        "—",
        "فعال",
        "—",
        "apps/inspection/application/services/add_inspection_item_service.py",
    ],
    [
        12,
        "راننده",
        "ارسال نهایی چک‌لیست",
        "راننده",
        "Submit چک‌لیست و تشخیص وجود آیتم خراب",
        "POST /api/v1/inspections/{id}/submit/",
        "status=SUBMITTED؛ خروجی has_failures",
        "خیر",
        "—",
        "خیر",
        "—",
        "inspection; inspection_item",
        "—",
        "—",
        "فعال",
        "FAIL به‌تنهایی Fault ایجاد نمی‌کند؛ باید مرحله گزارش خرابی اجرا شود.",
        "apps/inspection/application/services/submit_inspection_service.py",
    ],
    [
        13,
        "راننده",
        "خروج خودرو از مرکز",
        "راننده",
        "پس از Submit چک‌لیست و نبود Fault/Repair باز، خروج خودرو را ثبت می‌کند",
        "POST /api/v1/drivers/{id}/exit-center/",
        "vehicle.status=EXITED_CENTER",
        "خیر",
        "—",
        "خیر",
        "—",
        "vehicle",
        "—",
        "—",
        "فعال",
        "داده تخصیص از snapshot محلی SAP کنترل می‌شود.",
        "apps/driver/application/services/exit_center_service.py",
    ],
    [
        14,
        "خرابی",
        "گزارش خرابی از چک‌لیست",
        "راننده",
        "تبدیل آیتم‌های FAIL به یک Fault و اقلام خرابی",
        "POST /api/v1/inspections/{id}/report-fault/",
        "Fault در وضعیت OPEN؛ ارتباط با Inspection",
        "خیر",
        "کدهای عیب از fault_catalog محلی",
        "بله",
        "BAPI_ALM_NOTIF_CREATE",
        "fault; fault_item; sap_transaction",
        "fault.sap_notification_number",
        "SAP PM Notification",
        "فعال با Mock",
        "Write واقعی API v1 سیم‌کشی نشده؛ جزئیات در شیت نوشتن به SAP.",
        "apps/inspection/application/services/report_inspection_fault_service.py; infrastructure/sap/adapters/bapi/pm_notification_bapi_adapter.py",
    ],
    [
        15,
        "خرابی",
        "ارسال کیلومتر همراه خرابی چک‌لیست",
        "سیستم",
        "پس از موفقیت Notification، آخرین کیلومتر خودرو را به SAP می‌فرستد",
        "Side effect مرحله قبل",
        "Measurement Document و لاگ تراکنش",
        "خیر",
        "—",
        "بله",
        "MEASUREM_DOCUM_RFC_SINGLE_001",
        "sap_transaction؛ کیلومتر اصلی در vehicle_odometer_reading",
        "sap_transaction.sap_document_number",
        "SAP Measurement Document",
        "فعال با Mock",
        "فقط اگر Notification، measurement port و کیلومتر موجود باشند.",
        "apps/inspection/application/services/report_inspection_fault_service.py; infrastructure/sap/adapters/bapi/vehicle_measurement_bapi_adapter.py",
    ],
    [
        16,
        "خرابی",
        "گزارش خرابی دستی",
        "راننده / اپراتور",
        "ثبت یک یا چند خرابی مستقل از چک‌لیست",
        "POST /api/v1/faults/",
        "Fault/Items در وضعیت OPEN",
        "خیر",
        "کدها از fault_catalog محلی",
        "بله",
        "BAPI_ALM_NOTIF_CREATE",
        "fault; fault_item; sap_transaction",
        "fault.sap_notification_number",
        "SAP PM Notification",
        "فعال با Mock",
        "Write واقعی هنوز فعال نیست.",
        "apps/fault/application/services/report_fault_service.py; infrastructure/sap/adapters/bapi/pm_notification_bapi_adapter.py",
    ],
    [
        17,
        "خرابی",
        "ارسال کیلومتر همراه خرابی دستی",
        "سیستم",
        "پس از ایجاد Notification، آخرین کیلومتر را به SAP ارسال می‌کند",
        "Side effect ثبت خرابی",
        "Measurement Document و لاگ تراکنش",
        "خیر",
        "—",
        "بله",
        "MEASUREM_DOCUM_RFC_SINGLE_001",
        "sap_transaction؛ vehicle_odometer_reading",
        "sap_transaction.sap_document_number",
        "SAP Measurement Document",
        "فعال با Mock",
        "اگر کیلومتر محلی موجود نباشد این Write انجام نمی‌شود.",
        "apps/fault/application/services/report_fault_service.py; infrastructure/sap/adapters/bapi/vehicle_measurement_bapi_adapter.py",
    ],
    [
        18,
        "تصمیم توزیع",
        "خودرو قابل استفاده است",
        "سرپرست توزیع",
        "رد اثر عملیاتی خرابی و بازگرداندن خودرو به سرویس",
        "POST /api/v1/faults/{id}/distribution-usable/",
        "Fault بسته؛ Vehicle ACTIVE",
        "خیر",
        "—",
        "خیر",
        "—",
        "fault; vehicle",
        "fault.sap_notification_number حفظ می‌شود",
        "—",
        "فعال",
        "کد فعلی Notification SAP را Close نمی‌کند؛ در صورت نیاز BAPI را تأیید کنید.",
        "apps/fault/application/services/distribution_fault_decision_service.py; apps/fault/application/services/close_fault_service.py",
    ],
    [
        19,
        "تصمیم توزیع",
        "خودرو غیرقابل استفاده است",
        "سرپرست توزیع",
        "تایید توقف خودرو، ایجاد Repair Order و درخواست خودروی جایگزین",
        "POST /api/v1/faults/{id}/distribution-unusable/",
        "Fault=AWAITING_TRANSPORT؛ Vehicle=OUT_OF_SERVICE؛ Repair=CREATED",
        "خیر",
        "—",
        "بله",
        "ZFM_FLEET_ASSIGN_REPLACEMENT",
        "fault; vehicle; repair_order; repair_order_event; sap_transaction",
        "sap_transaction.sap_document_number",
        "SAP Vehicle Assignment (سفارشی)",
        "فعال با Mock / نام موقت",
        "نام Function سفارشی placeholder است و باید تیم SAP تأیید کند.",
        "apps/fault/application/services/distribution_fault_decision_service.py; infrastructure/sap/adapters/bapi/vehicle_assignment_bapi_adapter.py",
    ],
    [
        20,
        "ترابری",
        "تایید یا رد درخواست تعمیر",
        "سرپرست ترابری",
        "تایید Repair Order برای ادامه یا رد و بستن جریان",
        "POST /api/v1/repair-orders/{id}/approve/ یا transport-reject/",
        "تغییر وضعیت Repair/Fault/Vehicle و ثبت Timeline",
        "خیر",
        "—",
        "خیر",
        "—",
        "repair_order; repair_order_event; fault; vehicle",
        "—",
        "—",
        "فعال",
        "—",
        "apps/repair/application/services/approve_repair_order_service.py",
    ],
    [
        21,
        "ترابری",
        "تخصیص تعمیرگاه مرکزی/بیرونی",
        "سرپرست ترابری",
        "انتخاب نوع تعمیرگاه؛ برای بیرونی Referral ساخته می‌شود",
        "POST /api/v1/repair-orders/{id}/assign-workshop/",
        "Repair=WORKSHOP_ASSIGNED یا ایجاد Referral",
        "خیر",
        "—",
        "خیر",
        "—",
        "repair_order; external_workshop_referral_request; repair_order_event",
        "—",
        "—",
        "فعال",
        "هیچ سند SAP در این نقطه ساخته نمی‌شود.",
        "apps/repair/application/services/approve_repair_order_service.py",
    ],
    [
        22,
        "تعمیرگاه مرکزی",
        "تصمیم فنی: قابل تعمیر",
        "سرپرست تعمیرگاه",
        "پذیرش دستور تعمیر و شروع کار",
        "POST /api/v1/repair-orders/{id}/technical-decision/ یا accept/",
        "Repair=IN_PROGRESS؛ ایجاد PM Order SAP",
        "خیر",
        "—",
        "بله",
        "BAPI_ALM_ORDER_MAINTAIN",
        "repair_order; repair_order_event; sap_transaction",
        "repair_order.sap_order_number",
        "SAP PM Order",
        "فعال با Mock",
        "در UI فعلی Accept همین سرویس تصمیم فنی را فراخوانی می‌کند.",
        "apps/repair/application/services/workshop_technical_decision_service.py; apps/repair/application/services/sync_repair_to_sap_service.py",
    ],
    [
        23,
        "تعمیرگاه مرکزی",
        "تصمیم فنی: نیاز به تعمیر ندارد",
        "سرپرست تعمیرگاه",
        "رد فنی و پایان جریان تعمیر",
        "POST /api/v1/repair-orders/{id}/technical-decision/ یا reject/",
        "Repair=NO_REPAIR_NEEDED؛ Vehicle/Fault به وضعیت نهایی",
        "خیر",
        "—",
        "خیر",
        "—",
        "repair_order; repair_order_event; fault; vehicle",
        "—",
        "—",
        "فعال",
        "در صورت نیاز Close کردن PM Notification در SAP، BAPI باید مشخص/فعال شود.",
        "apps/repair/application/services/workshop_technical_decision_service.py",
    ],
    [
        24,
        "قطعات",
        "ثبت درخواست قطعه",
        "تکنسین / تعمیرگاه",
        "ثبت اقلام، مقدار و انتخاب از کاتالوگ/آزاد",
        "POST /api/v1/repair-orders/{id}/material-requests/",
        "Material Request=REQUESTED و snapshot موجودی هر قلم",
        "خیر",
        "central_stock محلی از ZI_STOCK_KH08_CDS",
        "خیر",
        "—",
        "material_request; material_request_item; repair_order_event",
        "—",
        "—",
        "فعال",
        "موجودی در لحظه از SAP خوانده نمی‌شود.",
        "apps/material/application/services/material_request_service.py",
    ],
    [
        25,
        "قطعات",
        "تصمیم تأمین از موجودی",
        "سرپرست ترابری",
        "انتخاب FROM_STOCK برای اقلام موجود",
        "POST /api/v1/material-requests/{id}/availability-decision/",
        "اقلام STOCK_ISSUED؛ ثبت تراکنش محلی",
        "خیر",
        "central_stock snapshot",
        "خیر در کد فعلی",
        "BAPI_GOODSMVT_CREATE_GI (Adapter موجود ولی وصل نیست)",
        "material_request; material_request_item; inventory_transaction",
        "—",
        "SAP Goods Issue مورد انتظار",
        "نیمه‌کاره",
        "برنامه فعلاً فقط local issue می‌زند؛ تعیین کنید GI باید همین‌جا به SAP ارسال شود یا خیر.",
        "apps/material/application/services/parts_availability_decision_service.py; infrastructure/sap/adapters/bapi/goods_issue_bapi_adapter.py",
    ],
    [
        26,
        "قطعات/خرید",
        "تصمیم خرید برای اقلام ناموجود",
        "سرپرست ترابری",
        "انتخاب PURCHASE؛ سیستم PR و line item می‌سازد",
        "POST /api/v1/material-requests/{id}/availability-decision/",
        "Material Request=PURCHASE_REQUIRED؛ PR=DRAFT",
        "خیر",
        "central_stock snapshot",
        "خیر",
        "—",
        "material_request; material_request_item; purchase_requisition; pr_line_item",
        "purchase_requisition.sap_pr_number بعداً",
        "—",
        "فعال",
        "ساخت PR در این مرحله فقط محلی است.",
        "apps/material/application/services/parts_availability_decision_service.py",
    ],
    [
        27,
        "خرید",
        "ایجاد/تکمیل دستی PR",
        "کارشناس خرید",
        "ایجاد PR یا افزودن line item",
        "POST /api/v1/purchase-requisitions/ و /{id}/line-items/",
        "PR و اقلام در FMMS",
        "خیر",
        "—",
        "خیر",
        "—",
        "purchase_requisition; pr_line_item",
        "—",
        "—",
        "فعال در Backend",
        "در فرانت فعلی صفحه اختصاصی Procurement دیده نمی‌شود.",
        "apps/procurement/application/services/create_purchase_requisition_service.py; apps/procurement/application/services/add_pr_line_item_service.py",
    ],
    [
        28,
        "خرید",
        "ارسال PR به SAP",
        "کارشناس خرید",
        "ارسال درخواست خرید محلی به SAP",
        "POST /api/v1/purchase-requisitions/{id}/submit-sap/",
        "PR=SUBMITTED؛ ذخیره شماره PR",
        "خیر",
        "—",
        "بله",
        "BAPI_PR_CREATE",
        "purchase_requisition; sap_transaction",
        "purchase_requisition.sap_pr_number",
        "SAP Purchase Requisition",
        "فعال با Mock",
        "Write واقعی API v1 هنوز سیم‌کشی نشده.",
        "apps/procurement/application/services/submit_pr_to_sap_service.py; infrastructure/sap/adapters/bapi/purchase_requisition_bapi_adapter.py",
    ],
    [
        29,
        "خرید",
        "دریافت PO از SAP/Caller",
        "SAP / سیستم واسط",
        "Payload سفارش خرید از Caller دریافت و در FMMS ثبت می‌شود",
        "POST /api/v1/purchase-orders/",
        "PO و line items محلی",
        "خیر",
        "Endpoint خودش SAP را نمی‌خواند",
        "خیر",
        "—",
        "purchase_order; po_line_item",
        "purchase_order.sap_po_number",
        "SAP PO موجود از قبل",
        "فعال در Backend",
        "روش ارسال payload از SAP/Middleware باید عملیاتی و مستند شود.",
        "apps/procurement/application/services/receive_po_from_sap_service.py; interfaces/api/v1/procurement/views.py",
    ],
    [
        30,
        "قطعات/خرید",
        "تحویل قطعه خریداری‌شده",
        "انبار / ترابری",
        "پس از خرید، اقلام را Issue شده اعلام می‌کند",
        "POST /api/v1/material-requests/{id}/issue-purchased/",
        "اقلام PURCHASED_ISSUED و تراکنش محلی",
        "خیر",
        "—",
        "خیر در کد فعلی",
        "نیازمند معرفی BAPI یا فعال‌سازی GI موجود",
        "material_request; material_request_item; inventory_transaction",
        "—",
        "SAP Goods Issue/Receipt احتمالی",
        "نیمه‌کاره",
        "مشخص شود GR و سپس GI لازم است یا فقط GI.",
        "apps/material/application/services/parts_availability_decision_service.py",
    ],
    [
        31,
        "تعمیرگاه مرکزی",
        "ثبت قطعات مصرفی و فعالیت‌ها",
        "تکنسین",
        "ثبت قطعه نصب‌شده، ساعت کار و توضیحات",
        "POST/PATCH/DELETE /api/v1/repair-orders/{id}/parts|activities/",
        "تاریخچه جزئیات تعمیر",
        "خیر",
        "—",
        "خیر",
        "—",
        "repair_part; repair_activity; repair_order_event",
        "repair_part.goods_issue_id (در کد فعلی غالباً خالی)",
        "—",
        "فعال",
        "ارتباط repair_part با سند GI SAP هنوز پیاده‌سازی عملی ندارد.",
        "apps/repair/application/services/add_repair_activity_service.py; apps/repair/infrastructure/models.py",
    ],
    [
        32,
        "تعمیرگاه مرکزی",
        "ثبت هزینه تعمیر داخلی",
        "سرپرست تعمیرگاه",
        "ثبت هزینه دستمزد، قطعه، خدمات و شماره فاکتور",
        "POST /api/v1/repair-orders/{id}/internal-cost/",
        "ثبت هزینه داخلی",
        "خیر",
        "—",
        "خیر",
        "—",
        "internal_repair_cost",
        "—",
        "—",
        "فعال",
        "سند مالی SAP ایجاد نمی‌شود.",
        "apps/repair/application/services/register_internal_repair_cost_service.py",
    ],
    [
        33,
        "تعمیرگاه مرکزی",
        "تکمیل فنی تعمیر",
        "تکنسین",
        "ثبت زمان تکمیل و ایجاد تحویل خودرو",
        "POST /api/v1/repair-orders/{id}/complete/",
        "Repair=WAITING_DRIVER_CONFIRMATION؛ Handover ساخته می‌شود",
        "خیر",
        "—",
        "خیر در کد فعلی",
        "BAPI_ALM_ORDER_COMPLETE (Adapter موجود ولی وصل نیست)",
        "repair_order; vehicle_handover; repair_order_event",
        "repair_order.sap_order_number",
        "SAP PM Order Completion مورد انتظار",
        "نیمه‌کاره",
        "تأیید کنید Order باید در این مرحله یا پس از تحویل راننده در SAP Complete شود.",
        "apps/repair/application/services/update_repair_status_service.py; infrastructure/sap/adapters/bapi/pm_order_bapi_adapter.py",
    ],
    [
        34,
        "تحویل",
        "تایید تحویل توسط راننده",
        "راننده",
        "قبول یا رد تعمیر انجام‌شده",
        "POST /api/v1/vehicle-handovers/{id}/confirm/",
        "قبول داخلی: Repair COMPLETED، Fault CLOSED، Vehicle ACTIVE؛ رد: Repair جدید",
        "خیر",
        "—",
        "خیر در کد فعلی",
        "نیازمند معرفی/تأیید BAPI بستن SAP",
        "vehicle_handover; repair_order; fault; vehicle; vehicle_component_history; repair_order_event",
        "repair_order.sap_order_number",
        "SAP PM Order/Notification Closure احتمالی",
        "فعال محلی",
        "بستن نهایی اسناد SAP در این نقطه پیاده‌سازی نشده است.",
        "apps/handover/application/services/handover_service.py",
    ],
    [
        35,
        "تعمیرگاه بیرونی",
        "ایجاد ارجاع و تخصیص تعمیرگاه",
        "ترابری",
        "ایجاد referral و سپس ثبت نام/آدرس/تاریخ تعمیرگاه بیرونی",
        "POST /api/v1/repair-orders/{id}/assign-external-workshop/",
        "Assignment و Timeline",
        "خیر",
        "—",
        "خیر در کد فعلی",
        "نیازمند معرفی BAPI ایجاد Service PO/PO",
        "external_workshop_referral_request; external_workshop_assignment; repair_order_event",
        "external_repair_review.sap_purchase_order_number بعداً",
        "SAP Service PO/PO مورد انتظار",
        "فعال محلی",
        "Adapterهای PO و Service PO موجودند ولی این workflow به آن‌ها وصل نیست.",
        "apps/repair/application/services/external_workshop_service.py; infrastructure/sap/adapters/bapi/service_po_bapi_adapter.py",
    ],
    [
        36,
        "تعمیرگاه بیرونی",
        "ثبت تحویل خودرو به تعمیرگاه",
        "ترابری",
        "ثبت تاریخ، کیلومتر، مشخصات تماس و یادداشت تحویل",
        "POST /api/v1/external-workshop-assignments/{id}/confirm-delivery/",
        "Assignment=DELIVERED",
        "خیر",
        "—",
        "خیر",
        "—",
        "external_workshop_delivery; external_workshop_assignment; vehicle; repair_order_event",
        "—",
        "—",
        "فعال",
        "کیلومتر این مرحله به SAP ارسال نمی‌شود.",
        "apps/repair/application/services/external_workshop_service.py",
    ],
    [
        37,
        "تعمیرگاه بیرونی",
        "ثبت دریافت خودرو از تعمیرگاه",
        "ترابری",
        "ثبت تاریخ و کیلومتر دریافت",
        "POST /api/v1/external-workshop-assignments/{id}/confirm-pickup/",
        "Assignment=PICKED_UP",
        "خیر",
        "—",
        "خیر",
        "—",
        "external_workshop_pickup; external_workshop_assignment; repair_order_event",
        "—",
        "—",
        "فعال",
        "—",
        "apps/repair/application/services/external_workshop_service.py",
    ],
    [
        38,
        "تعمیرگاه بیرونی",
        "بازبینی تعمیر و فاکتور",
        "ترابری",
        "آپلود فایل فاکتور، خدمات، قطعات تعویضی و هزینه",
        "POST /api/v1/external-workshop-assignments/{id}/review/",
        "Review در FMMS",
        "خیر",
        "—",
        "خیر در کد فعلی",
        "نیازمند معرفی BAPI Invoice Document",
        "external_repair_review؛ فایل در media/external_invoices/{assignment_id}/",
        "sap_purchase_order_number; sap_invoice_document_number (فعلاً خالی)",
        "SAP PO و Invoice Document مورد انتظار",
        "فعال محلی / SAP ناقص",
        "BAPI معادل برای Invoice و زمان ایجاد PO باید معرفی شود.",
        "apps/repair/application/services/external_workshop_service.py; interfaces/api/v1/repair/views.py",
    ],
    [
        39,
        "تعمیرگاه بیرونی",
        "بستن تعمیر بیرونی",
        "ترابری",
        "بستن assignment پس از بازبینی",
        "POST /api/v1/external-workshop-assignments/{id}/close/",
        "Assignment CLOSED؛ Repair/Fault/Component history به‌روزرسانی",
        "خیر",
        "—",
        "خیر در کد فعلی",
        "نیازمند معرفی/تأیید BAPI بستن SAP",
        "external_workshop_assignment; external_repair_review; repair_order; fault; vehicle_component_history; repair_order_event",
        "فیلدهای SAP در external_repair_review",
        "SAP PM/PO/Invoice Closure احتمالی",
        "فعال محلی",
        "چرخه اسناد SAP برای تعمیر بیرونی کامل نشده است.",
        "apps/repair/application/services/external_workshop_service.py",
    ],
    [
        40,
        "نگهداری پیشگیرانه",
        "ایجاد برنامه PM",
        "مدیر نگهداری",
        "تعریف interval، trigger و سررسید",
        "POST /api/v1/pm-plans/",
        "PM Plan محلی",
        "خیر",
        "—",
        "خیر",
        "—",
        "pm_plan",
        "—",
        "—",
        "فعال در Backend",
        "صفحه فرانت اختصاصی PM در کد فعلی دیده نمی‌شود.",
        "apps/preventive_maintenance/application/services/create_pm_plan_service.py",
    ],
    [
        41,
        "نگهداری پیشگیرانه",
        "Trigger دستور PM",
        "مدیر / زمان‌بند",
        "ایجاد PM Work Order و اعلان SAP",
        "POST /api/v1/pm-plans/{id}/trigger/",
        "PM Work Order=TRIGGERED؛ ذخیره شماره اعلان",
        "خیر",
        "—",
        "بله",
        "BAPI_ALM_NOTIF_CREATE",
        "pm_plan; pm_work_order; sap_transaction",
        "pm_work_order.sap_order_number (عملاً شماره Notification)",
        "SAP PM Notification",
        "فعال با Mock",
        "نام فیلد sap_order_number با نوع سند Notification همخوان نیست.",
        "apps/preventive_maintenance/application/services/trigger_pm_work_order_service.py",
    ],
    [
        42,
        "نگهداری پیشگیرانه",
        "تکمیل PM Work Order",
        "تکنسین",
        "ثبت تکمیل و یادداشت",
        "POST /api/v1/pm-work-orders/{id}/complete/",
        "PM Work Order=COMPLETED",
        "خیر",
        "—",
        "خیر در کد فعلی",
        "نیازمند معرفی/تأیید BAPI؛ Adapter PM Order Complete موجود است",
        "pm_work_order",
        "pm_work_order.sap_order_number",
        "SAP PM Notification/Order Closure",
        "فعال محلی",
        "نوع سند SAP و BAPI نهایی این مرحله باید مشخص شود.",
        "apps/preventive_maintenance/application/services/complete_pm_work_order_service.py",
    ],
    [
        43,
        "نظارت SAP",
        "مشاهده تراکنش‌های Write",
        "سرپرست / مدیر",
        "مشاهده وضعیت، payload، response، retry و شماره سند",
        "GET /api/v1/sap-transactions/",
        "Read از audit فنی FMMS",
        "خیر",
        "—",
        "خیر",
        "—",
        "sap_transaction",
        "sap_document_number",
        "—",
        "فعال",
        "برای بعضی object typeها retry adapter تعریف نشده است.",
        "apps/integration/infrastructure/models.py; interfaces/api/v1/integration/views.py",
    ],
    [
        44,
        "نظارت SAP",
        "Retry خودکار Writeهای ناموفق",
        "Celery",
        "هر ۱۵ دقیقه تراکنش‌های قابل retry را مجدداً ارسال می‌کند",
        "Task: fmms.retry_failed_sap_transactions",
        "status=RETRYING/SUCCESS/FAILED/EXHAUSTED",
        "خیر",
        "—",
        "بله",
        "بر اساس object type؛ PR/PM Order/Notification/Measurement/Assignment",
        "sap_transaction",
        "sap_document_number",
        "سند متناظر SAP",
        "فعال با Mock",
        "PO، GR، GI و Service PO در retry map فعلی نیستند.",
        "apps/integration/application/services/retry_failed_sap_transactions_service.py; infrastructure/messaging/tasks/sap_retry_tasks.py",
    ],
]


ODATA_HEADERS = [
    "دامنه",
    "وضعیت استفاده",
    "سرویس OData",
    "Entity / مسیر",
    "روش",
    "داده خوانده‌شده",
    "Trigger",
    "مقصد FMMS",
    "تنظیمات Env",
    "نکته",
    "منبع کد",
]

ODATA_ROWS = [
    [
        "خودرو و راننده",
        "فعال در Sync کلی",
        "ZC_VEHICLEDRIVER_CDS",
        "ZC_VehicleDriver",
        "GET JSON با pagination",
        "شماره خودرو، پلاک، تاریخ بهره‌برداری، راننده اول/دوم",
        "POST /sap-sync/ یا Celery",
        "vehicle; driver; vehicle_driver_assignment_history",
        "SAP_VEHICLE_DRIVER_SERVICE; SAP_VEHICLE_DRIVER_ENTITY_SET",
        "در config و compose پیش‌فرض Entity مشخص است.",
        "infrastructure/sap/adapters/odata/vehicle_driver_odata_adapter.py",
    ],
    [
        "قالب چک‌لیست",
        "فعال در Sync کلی",
        "ZI_FLEET_CAT_B_CDS",
        "root feed (entity خالی)",
        "GET XML",
        "CodeGroup, Code, GroupText, CodeText",
        "POST /sap-sync/ یا Celery",
        "inspection_template",
        "SAP_OBJECT_PART_CATALOG_SERVICE; SAP_OBJECT_PART_CATALOG_ENTITY_SET",
        "Adapter یک مسیر JSON قدیمی OBJECT_PART_CATALOG نیز دارد؛ Sync فعلی از XML استفاده می‌کند.",
        "infrastructure/sap/adapters/odata/object_part_catalog_odata_adapter.py",
    ],
    [
        "کاتالوگ خرابی",
        "فعال در Sync کلی",
        "ZI_B_DEFECTCATALOG9_CDS",
        "root feed (entity خالی)",
        "GET XML",
        "CodeGroup, Code, متن گروه/کد، کلاس خرابی",
        "POST /sap-sync/ یا Celery",
        "fault_catalog",
        "SAP_FAULT_CATALOG_SERVICE; SAP_FAULT_CATALOG_ENTITY_SET",
        "کاتالوگ در عملیات روزمره از دیتابیس محلی خوانده می‌شود.",
        "infrastructure/sap/adapters/odata/fault_catalog_odata_adapter.py",
    ],
    [
        "موجودی انبار مرکزی",
        "فعال در Sync کلی",
        "ZI_STOCK_KH08_CDS",
        "root feed (entity خالی)",
        "GET XML",
        "Material/Plant/SLoc/Quantity/Value/Currency",
        "POST /sap-sync/ یا Celery",
        "central_stock",
        "SAP_CENTRAL_STOCK_SERVICE; SAP_CENTRAL_STOCK_ENTITY_SET",
        "موجودی KH08 به‌صورت snapshot محلی مصرف می‌شود.",
        "infrastructure/sap/adapters/odata/central_stock_odata_adapter.py",
    ],
    [
        "Material Master",
        "Adapter موجود؛ استفاده‌نشده",
        "API_PRODUCT_SRV",
        "A_Product('{material}')؛ A_ProductPlant",
        "GET JSON",
        "مشخصات Material و Plant",
        "هیچ workflow فعلی",
        "هیچ مقصد فعالی در Sync کلی ندارد",
        "—",
        "در RunSAPSyncService سیم‌کشی نشده است.",
        "infrastructure/sap/adapters/odata/material_odata_adapter.py",
    ],
    [
        "موجودی استاندارد SAP",
        "Adapter موجود؛ استفاده‌نشده",
        "API_MATERIAL_STOCK_SRV",
        "MatlStkInAcctMod",
        "GET JSON",
        "Stock by material/plant",
        "هیچ workflow فعلی",
        "هیچ مقصد فعالی؛ تصمیم موجودی از central_stock است",
        "—",
        "در API v1 و Sync کلی سیم‌کشی نشده است.",
        "infrastructure/sap/adapters/odata/inventory_odata_adapter.py",
    ],
]


BAPI_HEADERS = [
    "رخداد / سند",
    "وضعیت",
    "Function Module",
    "زمان فراخوانی",
    "Object Type تراکنش",
    "کلید Idempotency",
    "ذخیره شماره برگشتی",
    "Adapter",
    "وضعیت اتصال واقعی",
    "اقدام لازم",
]

BAPI_ROWS = [
    [
        "ثبت خرابی دستی/چک‌لیست → PM Notification",
        "در workflow فعال",
        "BAPI_ALM_NOTIF_CREATE",
        "پس از ذخیره Fault",
        "FAULT",
        "fault-pm-notification:{fault_id}",
        "fault.sap_notification_number و sap_transaction.sap_document_number",
        "PMNotificationBAPIAdapter",
        "API v1 فقط Mock؛ real client سیم‌کشی نشده",
        "Composition root باید SAPBAPIClient واقعی را استفاده کند.",
    ],
    [
        "ثبت کیلومتر همراه Fault",
        "در workflow فعال",
        "MEASUREM_DOCUM_RFC_SINGLE_001",
        "بعد از Notification و در صورت وجود کیلومتر",
        "MEASUREMENT_DOCUMENT",
        "fault-odometer-measurement:{fault_id}",
        "sap_transaction.sap_document_number",
        "VehicleMeasurementBAPIAdapter",
        "API v1 فقط Mock",
        "Measurement point/contract با تیم SAP نهایی شود.",
    ],
    [
        "درخواست خودروی جایگزین",
        "در workflow فعال؛ نام موقت",
        "ZFM_FLEET_ASSIGN_REPLACEMENT",
        "پس از تصمیم distribution-unusable",
        "VEHICLE_ASSIGNMENT",
        "fault-replacement-assignment:{fault_id}",
        "sap_transaction.sap_document_number",
        "VehicleAssignmentBAPIAdapter",
        "API v1 فقط Mock",
        "نام و قرارداد Function سفارشی را تیم SAP تأیید کند.",
    ],
    [
        "ایجاد PM Order تعمیر",
        "در workflow فعال",
        "BAPI_ALM_ORDER_MAINTAIN",
        "در تصمیم فنی قابل تعمیر / sync-sap",
        "REPAIR_ORDER",
        "repair-pm-order:{repair_order_id}",
        "repair_order.sap_order_number و sap_transaction",
        "PMOrderBAPIAdapter",
        "API v1 فقط Mock",
        "کلاینت RFC واقعی سیم‌کشی شود.",
    ],
    [
        "ارسال PR",
        "در workflow فعال",
        "BAPI_PR_CREATE",
        "submit-sap",
        "PURCHASE_REQUISITION",
        "pr-submit:{pr_id} یا مقدار ورودی",
        "purchase_requisition.sap_pr_number و sap_transaction",
        "PurchaseRequisitionBAPIAdapter",
        "API v1 فقط Mock",
        "کلاینت RFC واقعی سیم‌کشی شود.",
    ],
    [
        "Trigger برنامه PM",
        "در workflow فعال",
        "BAPI_ALM_NOTIF_CREATE",
        "trigger PM plan",
        "PM_WORK_ORDER",
        "pm-notification:{work_order_id}",
        "pm_work_order.sap_order_number و sap_transaction",
        "PMNotificationBAPIAdapter",
        "API v1 فقط Mock",
        "تعیین شود فیلد محلی شماره Notification است یا Order.",
    ],
    [
        "بستن PM Notification",
        "Adapter موجود؛ workflow استفاده نمی‌کند",
        "BAPI_ALM_NOTIF_CLOSE",
        "نامشخص",
        "—",
        "—",
        "—",
        "PMNotificationBAPIAdapter",
        "غیرفعال",
        "مرحله مناسب بستن و کلید idempotency تعریف شود.",
    ],
    [
        "تکمیل PM Order",
        "Adapter موجود؛ workflow استفاده نمی‌کند",
        "BAPI_ALM_ORDER_COMPLETE",
        "نامشخص؛ احتمالاً پس از تحویل/تایید نهایی",
        "—",
        "—",
        "—",
        "PMOrderBAPIAdapter",
        "غیرفعال",
        "زمان بستن و نوع Order نهایی شود.",
    ],
    [
        "ایجاد/تایید PO",
        "Adapter موجود؛ workflow استفاده نمی‌کند",
        "BAPI_PO_CREATE1 / BAPI_PO_APPROVE",
        "نامشخص",
        "PURCHASE_ORDER بالقوه",
        "—",
        "purchase_order.sap_po_number بالقوه",
        "PurchaseOrderBAPIAdapter",
        "غیرفعال",
        "Workflow خرید فعلاً PO را فقط از caller دریافت می‌کند.",
    ],
    [
        "رسید/ابطال رسید کالا",
        "Adapter موجود؛ workflow استفاده نمی‌کند",
        "BAPI_GOODSMVT_CREATE_GR / BAPI_GOODSMVT_CANCEL_GR",
        "نامشخص",
        "GOODS_RECEIPT بالقوه",
        "—",
        "—",
        "GoodsReceiptBAPIAdapter",
        "غیرفعال",
        "ارتباط PO receipt با Material Request تعریف شود.",
    ],
    [
        "صدور/ابطال کالا",
        "Adapter موجود؛ workflow استفاده نمی‌کند",
        "BAPI_GOODSMVT_CREATE_GI / BAPI_GOODSMVT_CANCEL_GI",
        "هنگام issue از stock یا purchased",
        "GOODS_ISSUE بالقوه",
        "—",
        "repair_part.goods_issue_id بالقوه",
        "GoodsIssueBAPIAdapter",
        "غیرفعال",
        "در حال حاضر فقط inventory_transaction محلی ایجاد می‌شود.",
    ],
    [
        "Service PO تعمیرگاه بیرونی",
        "Adapter موجود؛ workflow استفاده نمی‌کند",
        "BAPI_SERVICE_PO_CREATE / BAPI_SERVICE_PO_CONFIRM",
        "ارجاع/تایید تعمیر بیرونی",
        "SERVICE_PO بالقوه",
        "—",
        "external_repair_review.sap_purchase_order_number بالقوه",
        "ServicePOBAPIAdapter",
        "غیرفعال",
        "تأیید شود این Functionها واقعی/سفارشی‌اند و در SAP مقصد وجود دارند.",
    ],
    [
        "ثبت Invoice تعمیرگاه بیرونی",
        "BAPI معادل در پروژه نیست",
        "نیازمند معرفی BAPI",
        "بازبینی/تایید فاکتور بیرونی",
        "—",
        "—",
        "external_repair_review.sap_invoice_document_number",
        "—",
        "پیاده‌سازی نشده",
        "BAPI یا API استاندارد Invoice Document توسط شما/تیم SAP معرفی شود.",
    ],
]


STORAGE_HEADERS = [
    "دامنه",
    "جدول / محل",
    "مدل",
    "چه چیزی ذخیره می‌شود",
    "منبع داده",
    "فیلدهای ارتباط با SAP",
    "ماندگاری / نکته",
    "منبع کد",
]

STORAGE_ROWS = [
    [
        "احراز هویت",
        "fmms_users",
        "FMMSUser",
        "کاربر، نقش، personnel number و وضعیت",
        "FMMS",
        "personnel_number برای تطبیق احتمالی",
        "PostgreSQL/SQLite طبق محیط",
        "apps/authentication/infrastructure/models.py",
    ],
    [
        "خودرو",
        "vehicle",
        "VehicleModel",
        "شماره خودرو، پلاک، تاریخ بهره‌برداری، راننده‌های تخصیص‌یافته و status",
        "SAP OData + تغییر وضعیت FMMS",
        "vehicle_number؛ driver customer numbers",
        "Soft-delete از BaseModel",
        "apps/vehicle/infrastructure/models.py",
    ],
    [
        "خودرو",
        "vehicle_driver_assignment_history",
        "VehicleDriverAssignmentHistoryModel",
        "Snapshot هر Sync از راننده اول/دوم",
        "SAP OData",
        "sync_run_id؛ vehicle_number؛ customer number",
        "تاریخچه append",
        "apps/vehicle/infrastructure/models.py",
    ],
    [
        "خودرو",
        "vehicle_odometer_reading",
        "VehicleOdometerReadingModel",
        "کیلومتر، تاریخ، منبع و ثبت‌کننده",
        "راننده/FMMS",
        "مبنای Measurement Document",
        "تاریخچه append",
        "apps/vehicle/infrastructure/models.py",
    ],
    [
        "خودرو",
        "vehicle_component_history",
        "VehicleComponentHistoryModel",
        "قطعات نصب‌شده پس از تعمیر",
        "FMMS Repair",
        "material_number؛ repair_order_id",
        "پس از تکمیل جریان ثبت می‌شود",
        "apps/vehicle/infrastructure/models.py",
    ],
    [
        "راننده",
        "driver",
        "DriverModel",
        "customer number، نام، موبایل، پرسنلی و status",
        "SAP OData",
        "customer_number؛ personnel_number",
        "Upsert در Sync",
        "apps/driver/infrastructure/models.py",
    ],
    [
        "بازرسی",
        "inspection",
        "InspectionModel",
        "سربرگ چک‌لیست، کیلومتر، نوع، زمان و status",
        "FMMS",
        "—",
        "رکورد عملیاتی",
        "apps/inspection/infrastructure/models.py",
    ],
    [
        "بازرسی",
        "inspection_item",
        "InspectionItemModel",
        "نتیجه و توضیح هر آیتم",
        "FMMS / قالب SAP",
        "item_id به قالب منطقی",
        "Child table",
        "apps/inspection/infrastructure/models.py",
    ],
    [
        "بازرسی",
        "inspection_template",
        "InspectionTemplateModel",
        "کدگروه/کد و متن آیتم‌های چک‌لیست",
        "SAP OData",
        "code_group; code",
        "Upsert/active flag",
        "apps/inspection/infrastructure/models.py",
    ],
    [
        "خرابی",
        "fault",
        "FaultModel",
        "خرابی، شدت، status، تصمیم توزیع و ارتباط با inspection",
        "FMMS + پاسخ SAP",
        "sap_defect_code; sap_notification_number",
        "رکورد اصلی خرابی",
        "apps/fault/infrastructure/models.py",
    ],
    [
        "خرابی",
        "fault_item",
        "FaultItemModel",
        "اجزای خرابی و شدت هر جزء",
        "FMMS/Inspection",
        "inspection_item_id",
        "Child logical by fault_id",
        "apps/fault/infrastructure/models.py",
    ],
    [
        "خرابی",
        "fault_catalog",
        "FaultCatalogModel",
        "کدهای خرابی، کلاس و توضیحات",
        "SAP OData",
        "code_group; code; defect_class",
        "Upsert/active flag",
        "apps/fault/infrastructure/models.py",
    ],
    [
        "تعمیر",
        "repair_order",
        "RepairOrderModel",
        "چرخه دستور تعمیر، تعمیرگاه، تکنسین و تاریخ تکمیل",
        "FMMS + پاسخ SAP",
        "sap_order_number",
        "هسته workflow تعمیر",
        "apps/repair/infrastructure/models.py",
    ],
    [
        "تعمیر",
        "repair_activity",
        "RepairActivityModel",
        "شرح فعالیت، ساعت کار و مجری",
        "FMMS",
        "—",
        "Child table",
        "apps/repair/infrastructure/models.py",
    ],
    [
        "تعمیر",
        "repair_part",
        "RepairPartModel",
        "قطعه مصرفی، مقدار و زمان ثبت",
        "FMMS",
        "goods_issue_id بالقوه",
        "اتصال GI فعلاً عملی نیست",
        "apps/repair/infrastructure/models.py",
    ],
    [
        "تعمیر",
        "repair_order_event",
        "RepairOrderEventModel",
        "Timeline رویدادها و actor",
        "FMMS",
        "—",
        "Audit کسب‌وکار",
        "apps/repair/infrastructure/models.py",
    ],
    [
        "تحویل",
        "vehicle_handover",
        "VehicleHandoverModel",
        "قبول/رد تحویل، راننده، نظر و زمان",
        "FMMS",
        "—",
        "یک رکورد یکتا برای repair_order",
        "apps/handover/infrastructure/models.py",
    ],
    [
        "قطعات",
        "material_request",
        "MaterialRequestModel",
        "درخواست قطعه برای Repair Order و status",
        "FMMS",
        "—",
        "Parent request",
        "apps/material/infrastructure/models.py",
    ],
    [
        "قطعات",
        "material_request_item",
        "MaterialRequestItemModel",
        "ماده، مقدار، تصمیم تأمین و snapshot موجودی",
        "FMMS + snapshot SAP",
        "material_number",
        "Child table",
        "apps/material/infrastructure/models.py",
    ],
    [
        "قطعات",
        "inventory_transaction",
        "InventoryTransactionModel",
        "صدور موجودی محلی",
        "FMMS",
        "material_request_id",
        "سند SAP GI نیست",
        "apps/material/infrastructure/models.py",
    ],
    [
        "قطعات",
        "central_stock",
        "CentralStockModel",
        "موجودی و ارزش انبار مرکزی",
        "SAP OData",
        "material; plant; storage_location",
        "Snapshot محلی/active flag",
        "apps/material/infrastructure/models.py",
    ],
    [
        "خرید",
        "purchase_requisition",
        "PurchaseRequisitionModel",
        "PR، status و ارتباط با material request",
        "FMMS + پاسخ SAP",
        "sap_pr_number",
        "شماره SAP بعد از submit",
        "apps/procurement/infrastructure/models.py",
    ],
    [
        "خرید",
        "pr_line_item",
        "PRLineItemModel",
        "اقلام PR، مقدار و قیمت تخمینی",
        "FMMS",
        "material_number",
        "Child table",
        "apps/procurement/infrastructure/models.py",
    ],
    [
        "خرید",
        "purchase_order",
        "PurchaseOrderModel",
        "PO دریافت‌شده از caller/SAP",
        "Inbound payload",
        "sap_po_number; vendor_number",
        "Endpoint خودش SAP را صدا نمی‌زند",
        "apps/procurement/infrastructure/models.py",
    ],
    [
        "خرید",
        "po_line_item",
        "POLineItemModel",
        "اقلام PO و مقدار دریافت‌شده",
        "Inbound payload",
        "material_number",
        "Child table",
        "apps/procurement/infrastructure/models.py",
    ],
    [
        "PM",
        "pm_plan",
        "PMPlanModel",
        "برنامه نگهداری، interval و trigger",
        "FMMS",
        "—",
        "تعریف محلی",
        "apps/preventive_maintenance/infrastructure/models.py",
    ],
    [
        "PM",
        "pm_work_order",
        "PMWorkOrderModel",
        "دستور PM، زمان‌ها، status و شماره سند SAP",
        "FMMS + پاسخ SAP",
        "sap_order_number",
        "در Trigger فعلی شماره Notification نگهداری می‌شود",
        "apps/preventive_maintenance/infrastructure/models.py",
    ],
    [
        "یکپارچه‌سازی",
        "sap_sync_run",
        "SAPSyncRunModel",
        "اجرای کلی Sync، trigger، خلاصه و خطا",
        "FMMS Integration",
        "request_id",
        "Audit خواندن SAP",
        "apps/integration/infrastructure/models.py",
    ],
    [
        "یکپارچه‌سازی",
        "sap_sync_run_item",
        "SAPSyncRunItemModel",
        "نتیجه هر زیر Sync",
        "FMMS Integration",
        "sync_run FK",
        "Audit خواندن SAP",
        "apps/integration/infrastructure/models.py",
    ],
    [
        "یکپارچه‌سازی",
        "sap_transaction",
        "SAPTransactionModel",
        "payload/response، retry، idempotency، status و شماره سند",
        "همه Writeهای SAP",
        "object_type; sap_document_number",
        "Audit و retry فنی",
        "apps/integration/infrastructure/models.py",
    ],
    [
        "تعمیر بیرونی",
        "external_workshop_referral_request",
        "ExternalWorkshopReferralRequestModel",
        "درخواست ارجاع و تایید/رد",
        "FMMS",
        "—",
        "Workflow محلی",
        "apps/repair/infrastructure/models.py",
    ],
    [
        "تعمیر بیرونی",
        "external_workshop_assignment",
        "ExternalWorkshopAssignmentModel",
        "تخصیص تعمیرگاه، وضعیت و لغو",
        "FMMS",
        "—",
        "Workflow محلی",
        "apps/repair/infrastructure/models.py",
    ],
    [
        "تعمیر بیرونی",
        "external_workshop_delivery",
        "ExternalWorkshopDeliveryModel",
        "اطلاعات تحویل و کیلومتر",
        "FMMS",
        "—",
        "One-to-one با assignment",
        "apps/repair/infrastructure/models.py",
    ],
    [
        "تعمیر بیرونی",
        "external_workshop_pickup",
        "ExternalWorkshopPickupModel",
        "اطلاعات دریافت و کیلومتر",
        "FMMS",
        "—",
        "One-to-one با assignment",
        "apps/repair/infrastructure/models.py",
    ],
    [
        "تعمیر بیرونی",
        "external_repair_review",
        "ExternalRepairReviewModel",
        "خدمات، قطعات، هزینه، فایل فاکتور و شماره‌های SAP",
        "FMMS؛ SAP write هنوز ندارد",
        "sap_purchase_order_number; sap_invoice_document_number",
        "فیلدهای SAP فعلاً توسط workflow پر نمی‌شوند",
        "apps/repair/infrastructure/models.py",
    ],
    [
        "تعمیر بیرونی",
        "external_repair_invoice",
        "ExternalRepairInvoiceModel",
        "فاکتور بیرونی، مبلغ، ارز و تایید",
        "FMMS",
        "—",
        "مسیر invoice قدیمی/مجزا",
        "apps/repair/infrastructure/models.py",
    ],
    [
        "تعمیر داخلی",
        "internal_repair_cost",
        "InternalRepairCostModel",
        "هزینه‌های داخلی و شماره فاکتور",
        "FMMS",
        "—",
        "سند مالی SAP ندارد",
        "apps/repair/infrastructure/models.py",
    ],
    [
        "فایل",
        "media/external_invoices/{assignment_id}/",
        "FileSystemStorage",
        "فایل فاکتور آپلودشده تعمیرگاه بیرونی",
        "کاربر",
        "مسیر در external_repair_review.invoice_attachment",
        "فایل خارج از DB؛ فقط path در DB",
        "interfaces/api/v1/repair/views.py",
    ],
    [
        "مرورگر",
        "localStorage",
        "Web Storage",
        "JWT access/refresh و زمان انقضا",
        "Frontend",
        "—",
        "روی مرورگر کاربر؛ نه در DB سرور",
        "frontend/src/api/client.ts",
    ],
]


GAP_HEADERS = [
    "اولویت",
    "موضوع",
    "وضعیت فعلی",
    "تصمیم / اطلاعات موردنیاز",
    "اثر",
    "منبع",
]

GAP_ROWS = [
    [
        "خیلی بالا",
        "اتصال واقعی BAPI",
        "_sap_client فقط MockSAPClient می‌سازد و با SAP_USE_MOCK=False خطا می‌دهد.",
        "SAPBAPIClient در composition root سیم‌کشی و credential/connection تست شود.",
        "تمام Writeهای SAP فعلاً واقعی نیستند.",
        "interfaces/api/v1/deps.py",
    ],
    [
        "خیلی بالا",
        "BAPI خودروی جایگزین",
        "ZFM_FLEET_ASSIGN_REPLACEMENT placeholder است.",
        "نام و signature نهایی Function سفارشی را تیم SAP اعلام کند.",
        "مرحله distribution-unusable وابسته است.",
        "infrastructure/sap/adapters/bapi/vehicle_assignment_bapi_adapter.py",
    ],
    [
        "بالا",
        "صدور قطعه از انبار",
        "فقط inventory_transaction محلی ثبت می‌شود.",
        "تأیید شود BAPI_GOODSMVT_CREATE_GI در مرحله FROM_STOCK و PURCHASED اجرا شود.",
        "موجودی SAP ممکن است با FMMS مغایر شود.",
        "apps/material/application/services/parts_availability_decision_service.py",
    ],
    [
        "بالا",
        "رسید قطعه خریداری‌شده",
        "GR در workflow فعال نیست.",
        "تعیین شود دریافت PO/قطعه با BAPI_GOODSMVT_CREATE_GR انجام شود و کجا ذخیره گردد.",
        "زنجیره PR→PO→GR→GI ناقص است.",
        "infrastructure/sap/adapters/bapi/goods_receipt_bapi_adapter.py",
    ],
    [
        "بالا",
        "بستن PM Order/Notification",
        "Completion محلی است؛ Adapterهای close/complete استفاده نمی‌شوند.",
        "تعیین زمان دقیق و BAPI نهایی بستن سند SAP.",
        "اسناد SAP ممکن است باز بمانند.",
        "infrastructure/sap/adapters/bapi/pm_order_bapi_adapter.py",
    ],
    [
        "بالا",
        "تعمیرگاه بیرونی و Service PO",
        "فیلد PO وجود دارد اما SAP write ندارد.",
        "تأیید BAPI/Service API ایجاد PO و زمان فراخوانی.",
        "هزینه بیرونی به سند خرید SAP متصل نیست.",
        "apps/repair/infrastructure/models.py",
    ],
    [
        "بالا",
        "Invoice Document بیرونی",
        "هیچ BAPI معادل در پروژه نیست.",
        "BAPI یا API ایجاد Invoice Document معرفی شود.",
        "sap_invoice_document_number خالی می‌ماند.",
        "apps/repair/application/services/external_workshop_service.py",
    ],
    [
        "متوسط",
        "PO ورودی",
        "POST /purchase-orders payload را از caller می‌گیرد و SAP را صدا نمی‌زند.",
        "Middleware/مالک trigger، schema و احراز هویت مشخص شود.",
        "دریافت PO عملیاتی مبهم است.",
        "apps/procurement/application/services/receive_po_from_sap_service.py",
    ],
    [
        "متوسط",
        "نوع شماره PM Work Order",
        "Trigger PM notification می‌سازد ولی در sap_order_number ذخیره می‌کند.",
        "نام فیلد یا نوع سند اصلاح/مستند شود.",
        "گزارش‌گیری ممکن است شماره Notification را Order تلقی کند.",
        "apps/preventive_maintenance/application/services/trigger_pm_work_order_service.py",
    ],
    [
        "متوسط",
        "Retry Adapterها",
        "PO/GR/GI/Service PO در retry map نیستند.",
        "پس از فعال‌سازی workflow، adapterهای retry اضافه شوند.",
        "خطاهای SAP این اسناد خودکار جبران نمی‌شوند.",
        "apps/integration/application/services/retry_failed_sap_transactions_service.py",
    ],
    [
        "متوسط",
        "دو نام‌گذاری Env",
        "SAPConfig و settings برخی نام‌های متفاوت SAP_* دارند.",
        "متغیرها یکدست شوند.",
        "ریسک خطای deployment.",
        "infrastructure/sap/config.py; config/settings/base.py",
    ],
]


def style_table_sheet(ws, headers, rows, table_name, widths):
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).column_letter}{len(rows) + 1}"
    table = Table(
        displayName=table_name,
        ref=f"A1:{ws.cell(1, len(headers)).column_letter}{len(rows) + 1}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(bottom=Side(style="medium", color=TEAL))
    ws.row_dimensions[1].height = 38
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=9, color=DARK)
            cell.alignment = Alignment(
                horizontal="right", vertical="top", wrap_text=True
            )
            cell.border = Border(bottom=LIGHT_BORDER)
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 58


def build_summary(wb):
    ws = wb.active
    ws.title = "راهنما و خلاصه"
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    ws["A1"] = "نقشه جامع فرآیند FMMS و یکپارچگی SAP"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A4:H4")
    ws["A4"] = (
        "مبنای تهیه: کد فعلی پروژه در تاریخ 2026-07-27. "
        "OData یعنی Read از SAP؛ BAPI/RFC یعنی Write به SAP. "
        "هیچ credential یا URL حساس در فایل درج نشده است."
    )
    ws["A4"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A4"].font = Font(name="Arial", size=10, color=DARK)
    ws["A4"].alignment = Alignment(
        horizontal="right", vertical="center", wrap_text=True
    )
    ws.row_dimensions[4].height = 42

    labels = [
        ("A6:B6", "تعداد مراحل مستند", "=COUNTA('فرآیند انتها به انتها'!A2:A200)"),
        (
            "C6:D6",
            "مراحل دارای SAP Write",
            "=COUNTIF('فرآیند انتها به انتها'!J2:J200,\"بله\")",
        ),
        ("E6:F6", "ODataهای فعال", "=COUNTIF('خواندن از SAP'!B2:B100,\"فعال*\")"),
        (
            "G6:H6",
            "شکاف‌های خیلی مهم",
            "=COUNTIF('شکاف‌ها و سوالات'!A2:A100,\"خیلی بالا\")",
        ),
    ]
    for merged, label, formula in labels:
        start, end = merged.split(":")
        start_col = ws[start].column
        end_col = ws[end].column
        row = ws[start].row
        ws.merge_cells(
            start_row=row, start_column=start_col, end_row=row, end_column=end_col
        )
        ws[start] = label
        ws[start].fill = PatternFill("solid", fgColor=TEAL)
        ws[start].font = Font(name="Arial", bold=True, color=WHITE, size=10)
        ws[start].alignment = Alignment(horizontal="center")
        value_row = row + 1
        ws.merge_cells(
            start_row=value_row,
            start_column=start_col,
            end_row=value_row,
            end_column=end_col,
        )
        value_cell = ws.cell(value_row, start_col)
        value_cell.value = formula
        value_cell.fill = PatternFill("solid", fgColor="EAF4F4")
        value_cell.font = Font(name="Arial", bold=True, color=NAVY, size=18)
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.border = Border(bottom=Side(style="medium", color=TEAL))
    ws.row_dimensions[6].height = 26
    ws.row_dimensions[7].height = 34

    ws.merge_cells("A10:H10")
    ws["A10"] = "نتیجه کلیدی ممیزی"
    ws["A10"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A10"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws["A10"].alignment = Alignment(horizontal="center")
    notes = [
        "۱) چهار Read فعال SAP از طریق OData به snapshotهای محلی منتقل می‌شوند: خودرو/راننده، قالب چک‌لیست، کاتالوگ خرابی و موجودی انبار مرکزی.",
        "۲) Writeهای تعریف‌شده از SAPTransactionManager عبور می‌کنند و payload/response/idempotency در sap_transaction ذخیره می‌شود.",
        "۳) در ترکیب فعلی API v1، BAPIها فقط با MockSAPClient اجرا می‌شوند؛ Write واقعی به SAP عملیاتی نیست.",
        "۴) برای Goods Issue/Receipt، بستن PM Order/Notification و اسناد تعمیرگاه بیرونی Adapter یا فیلد وجود دارد، اما workflow کامل نیست.",
        "۵) هر سطر شیت فرآیند مشخص می‌کند Read مستقیم SAP دارد یا از snapshot محلی استفاده می‌کند، Write دارد یا نه، و داده دقیقاً کجا ذخیره می‌شود.",
    ]
    for idx, note in enumerate(notes, 11):
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)
        ws.cell(idx, 1).value = note
        ws.cell(idx, 1).font = Font(name="Arial", size=10, color=DARK)
        ws.cell(idx, 1).alignment = Alignment(
            horizontal="right", vertical="center", wrap_text=True
        )
        ws.cell(idx, 1).fill = PatternFill(
            "solid", fgColor=RED if idx == 13 else ("FFF8E1" if idx in {14} else WHITE)
        )
        ws.row_dimensions[idx].height = 32

    ws.merge_cells("A18:H18")
    ws["A18"] = "راهنمای رنگ‌ها و استفاده"
    ws["A18"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A18"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws["A18"].alignment = Alignment(horizontal="center")
    guide = [
        ["سبز", "فعال/پیاده‌سازی‌شده", GREEN, "برای عملیات موجود در کد فعلی"],
        [
            "زرد",
            "نیمه‌کاره یا نیازمند تصمیم",
            YELLOW,
            "Adapter/فیلد وجود دارد ولی workflow کامل نیست",
        ],
        ["قرمز", "ریسک یا اتصال واقعی غیرفعال", RED, "مانند Mock بودن Writeهای SAP"],
        ["خاکستری", "بدون ارتباط مستقیم SAP", GRAY, "عملیات صرفاً محلی FMMS"],
    ]
    for i, (color_name, meaning, color, usage) in enumerate(guide, 19):
        ws[f"A{i}"] = color_name
        ws[f"B{i}"] = meaning
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=8)
        ws[f"C{i}"] = usage
        for cell in ws[i]:
            cell.font = Font(name="Arial", size=9, color=DARK)
            cell.alignment = Alignment(
                horizontal="right", vertical="center", wrap_text=True
            )
        ws[f"A{i}"].fill = PatternFill("solid", fgColor=color)
        ws[f"B{i}"].fill = PatternFill("solid", fgColor=color)
        ws[f"C{i}"].fill = PatternFill("solid", fgColor=color)
        ws.row_dimensions[i].height = 26

    for col, width in zip("ABCDEFGH", [14, 18, 14, 18, 14, 18, 14, 18], strict=True):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"


def apply_process_formatting(ws):
    max_row = ws.max_row
    ws.conditional_formatting.add(
        f"J2:J{max_row}",
        FormulaRule(formula=['$J2="بله"'], fill=PatternFill("solid", fgColor=ORANGE)),
    )
    ws.conditional_formatting.add(
        f"O2:O{max_row}",
        FormulaRule(
            formula=['ISNUMBER(SEARCH("فعال",$O2))'],
            fill=PatternFill("solid", fgColor=GREEN),
        ),
    )
    ws.conditional_formatting.add(
        f"O2:O{max_row}",
        FormulaRule(
            formula=['OR(ISNUMBER(SEARCH("نیمه",$O2)),ISNUMBER(SEARCH("ناقص",$O2)))'],
            fill=PatternFill("solid", fgColor=YELLOW),
        ),
    )
    ws.conditional_formatting.add(
        f"O2:O{max_row}",
        FormulaRule(
            formula=['ISNUMBER(SEARCH("Mock",$O2))'],
            fill=PatternFill("solid", fgColor=RED),
        ),
    )
    dv = DataValidation(
        type="list",
        formula1='"بله,خیر,خیر در کد فعلی"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"J2:J{max_row}")


def build_workbook():
    wb = Workbook()
    build_summary(wb)

    ws = wb.create_sheet("فرآیند انتها به انتها")
    style_table_sheet(
        ws,
        PROCESS_HEADERS,
        PROCESS_ROWS,
        "EndToEndProcess",
        [7, 18, 25, 18, 42, 34, 38, 16, 32, 14, 33, 38, 28, 30, 24, 42, 48],
    )
    apply_process_formatting(ws)

    ws = wb.create_sheet("خواندن از SAP")
    style_table_sheet(
        ws,
        ODATA_HEADERS,
        ODATA_ROWS,
        "SAPReads",
        [20, 22, 27, 28, 18, 38, 26, 38, 38, 44, 52],
    )
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 76

    ws = wb.create_sheet("نوشتن به SAP")
    style_table_sheet(
        ws,
        BAPI_HEADERS,
        BAPI_ROWS,
        "SAPWrites",
        [34, 27, 38, 34, 25, 38, 42, 32, 34, 44],
    )
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 82
    ws.conditional_formatting.add(
        f"I2:I{ws.max_row}",
        FormulaRule(
            formula=['ISNUMBER(SEARCH("Mock",$I2))'],
            fill=PatternFill("solid", fgColor=RED),
        ),
    )

    ws = wb.create_sheet("محل ذخیره داده")
    style_table_sheet(
        ws,
        STORAGE_HEADERS,
        STORAGE_ROWS,
        "DataStorage",
        [20, 38, 32, 48, 29, 44, 38, 55],
    )
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 72

    ws = wb.create_sheet("شکاف‌ها و سوالات")
    style_table_sheet(
        ws,
        GAP_HEADERS,
        GAP_ROWS,
        "GapsQuestions",
        [15, 35, 53, 55, 42, 55],
    )
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 88
    ws.conditional_formatting.add(
        f"A2:A{ws.max_row}",
        FormulaRule(
            formula=['$A2="خیلی بالا"'], fill=PatternFill("solid", fgColor=RED)
        ),
    )
    ws.conditional_formatting.add(
        f"A2:A{ws.max_row}",
        FormulaRule(formula=['$A2="بالا"'], fill=PatternFill("solid", fgColor=ORANGE)),
    )
    ws.conditional_formatting.add(
        f"A2:A{ws.max_row}",
        FormulaRule(formula=['$A2="متوسط"'], fill=PatternFill("solid", fgColor=YELLOW)),
    )

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    inject_formula_cache()


def inject_formula_cache():
    """Add cached KPI values while preserving formulas for non-calculating previews."""
    cached_values = {
        "A7": len(PROCESS_ROWS),
        "C7": sum(row[9] == "بله" for row in PROCESS_ROWS),
        "E7": sum(row[1].startswith("فعال") for row in ODATA_ROWS),
        "G7": sum(row[0] == "خیلی بالا" for row in GAP_ROWS),
    }
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        with zipfile.ZipFile(OUTPUT_FILE, "r") as source_zip:
            source_zip.extractall(tmp_path)
        sheet_xml = tmp_path / "xl" / "worksheets" / "sheet1.xml"
        xml = sheet_xml.read_text(encoding="utf-8")
        for coordinate, value in cached_values.items():
            pattern = rf'(<c r="{coordinate}"[^>]*><f>.*?</f>)' rf"<v\s*/>" rf"(</c>)"
            xml, replacements = re.subn(
                pattern,
                rf"\g<1><v>{value}</v>\g<2>",
                xml,
                count=1,
            )
            assert replacements == 1, f"Formula cache cell {coordinate} not found."
        sheet_xml.write_text(xml, encoding="utf-8")
        rebuilt = OUTPUT_FILE.with_suffix(".rebuilt.xlsx")
        with zipfile.ZipFile(rebuilt, "w", zipfile.ZIP_DEFLATED) as target_zip:
            for path in sorted(tmp_path.rglob("*")):
                if path.is_file():
                    target_zip.write(path, path.relative_to(tmp_path))
        shutil.move(rebuilt, OUTPUT_FILE)


def create_qa_active_sheet_copies():
    """Create temporary visual-QA copies with each worksheet first and active."""
    qa_dir = OUTPUT_DIR / "qa_active_sheets"
    qa_dir.mkdir(parents=True, exist_ok=True)
    paths = []
    sheet_names = load_workbook(OUTPUT_FILE, read_only=True).sheetnames
    for index, sheet_name in enumerate(sheet_names):
        workbook = load_workbook(OUTPUT_FILE, data_only=False)
        target = workbook[sheet_name]
        workbook._sheets.remove(target)
        workbook._sheets.insert(0, target)
        workbook.active = 0
        safe_name = f"{index + 1:02d}.xlsx"
        path = qa_dir / safe_name
        workbook.save(path)
        paths.append(str(path))
    return paths


def verify_workbook():
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    expected_sheets = [
        "راهنما و خلاصه",
        "فرآیند انتها به انتها",
        "خواندن از SAP",
        "نوشتن به SAP",
        "محل ذخیره داده",
        "شکاف‌ها و سوالات",
    ]
    assert wb.sheetnames == expected_sheets
    assert wb["فرآیند انتها به انتها"].max_row == len(PROCESS_ROWS) + 1
    assert wb["خواندن از SAP"].max_row == len(ODATA_ROWS) + 1
    assert wb["نوشتن به SAP"].max_row == len(BAPI_ROWS) + 1
    assert wb["محل ذخیره داده"].max_row == len(STORAGE_ROWS) + 1
    assert wb["شکاف‌ها و سوالات"].max_row == len(GAP_ROWS) + 1
    formulas = []
    errors = []
    for ws in wb.worksheets:
        assert ws.sheet_view.rightToLeft
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append((ws.title, cell.coordinate, cell.value))
                if isinstance(cell.value, str) and any(
                    marker in cell.value
                    for marker in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                ):
                    errors.append((ws.title, cell.coordinate, cell.value))
    assert len(formulas) == 4
    assert not errors
    return {
        "path": str(OUTPUT_FILE),
        "sheets": wb.sheetnames,
        "process_rows": len(PROCESS_ROWS),
        "odata_rows": len(ODATA_ROWS),
        "bapi_rows": len(BAPI_ROWS),
        "storage_rows": len(STORAGE_ROWS),
        "gap_rows": len(GAP_ROWS),
        "formulas": formulas,
    }


if __name__ == "__main__":
    build_workbook()
    result = verify_workbook()
    result["qa_files"] = create_qa_active_sheet_copies()
    print(result)  # noqa: T201 - intentional CLI output
